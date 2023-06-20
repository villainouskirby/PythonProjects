from tkinter import *
from tkinter import filedialog
import tkinter.ttk as ttk
import tkinter.messagebox as msgbox
import pandas as pd
from openpyxl import *
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side
import os
import sys

def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.abspath(".")

    return os.path.join(base_path, relative_path)

title_path = sys.argv[0]
title = os.path.splitext(os.path.basename(title_path))

root = Tk()
root.title(title[0])

def need_help():
    global img_idx
    h = Toplevel()
    h.title("도움말")
    h.geometry("400x520")

    img_idx = 0

    def h_next():
        global img_idx
        img_idx += 1
        h_update()

    def h_prev():
        global img_idx
        img_idx -= 1
        h_update()

    def h_update():
        if img_idx == 3:
            n_btn.config(state=DISABLED)
        else:
            n_btn.config(state=NORMAL)
        if img_idx == 0:
            p_btn.config(state=DISABLED)
        else:
            p_btn.config(state=NORMAL)

        img_label.config(image=img_lst[img_idx])
        txt_label.config(text=txt_lst[img_idx])

    img_frame = Frame(h)
    img_frame.pack(fill="x", padx=5, pady=5, ipady=5)

    img_label = Label(img_frame)
    img_label.pack(padx=5, pady=5)

    txt_label = Label(img_frame)
    txt_label.pack(padx=5, pady=5)

    h_btn_frame = Frame(h)
    h_btn_frame.pack(fill="x", padx=5, pady=5, ipady=5, side="bottom")

    n_btn = switch_button(h_btn_frame, "다음", h_next)
    n_btn.pack(side="right", padx=5, pady=5)

    p_btn = switch_button(h_btn_frame, "이전", h_prev)
    p_btn.pack(side="right", padx=5, pady=5)

    h_exit_btn = switch_button(h_btn_frame, "닫기", h.destroy)
    h_exit_btn.pack(side="left", padx=5, pady=5)

    h_update()

    h.iconbitmap(resource_path("project_file/kirby.ico"))
    h.resizable(False, False)
    h.mainloop()

def switch_button(frame, txt, cmd):
    return Button(frame, width=12, padx=5, pady=5, text=txt, command=cmd)

def switch_frame(frame_class):
    for i in [page1, page2, page3, page4]:
        i.pack_forget()
    frame_class.pack(fill="both", expand=True)

def open_file():
    file_selected = filedialog.askopenfilename(title="열기",filetypes=(("xlsx 파일", "*.xlsx"), ("모든 파일", "*.*")))
    if file_selected == txt_file_dest.get():
        return
    if file_selected == '':
        return

    txt_file_dest.config(state="normal")
    txt_file_dest.delete(0, END)
    txt_file_dest.insert(0, file_selected)
    txt_file_dest.config(state="readonly")

    sheet_box.config(values=load_workbook(txt_file_dest.get()).get_sheet_names())
    sheet_box.current(0)

def data_arrange(a):
    if a in data_idx_lst:
        return
    data_lst.append(db_col[a])
    data_idx_lst.append(a)
    data_update()

def data_update():
    data_label.config(text=data_selected + ', '.join(data_lst))

def data_erase():
    data_lst.pop()
    data_idx_lst.pop()
    data_update()

def data_erase_all():
    data_lst.clear()
    data_idx_lst.clear()
    data_update()

def check():
    global db, db_col, treeview, scrollbar, db_list
    if txt_file_dest.get() == '':
        msgbox.showwarning("경고", "파일을 추가하십시오.")
        return

    db = pd.read_excel(txt_file_dest.get(), sheet_name=sheet_box.get(), index_col=None).fillna(0)
    db_col = list([i for i in db])
    db_list = []
    for i in db.values.tolist():
        temp_lst = []
        for j in i:
            if type(j) == float:
                temp_lst.append(int(j))
            else:
                temp_lst.append(j)
        db_list.append(temp_lst)

    scrollbar = Scrollbar(whole_data)
    scrollbar.pack(side="right", fill="y")

    treeview = ttk.Treeview(whole_data, columns=db_col, displaycolumns=db_col,show="headings", yscrollcommand = scrollbar.set)
    treeview.pack(fill='both', expand=True)
    for i in db_col:
        treeview.column(i, width=50, anchor="center")
        treeview.heading(i, text=i, anchor="center")
    for i in db_list:
        treeview.insert('', 'end', values=i)
    l = len(db_col)
    if l > 8:
        treeview.heading(db_col[8], command=lambda:data_arrange(8))
    if l > 7:
        treeview.heading(db_col[7], command=lambda:data_arrange(7))
    if l > 6:
        treeview.heading(db_col[6], command=lambda:data_arrange(6))
    if l > 5:
        treeview.heading(db_col[5], command=lambda:data_arrange(5))
    if l > 4:
        treeview.heading(db_col[4], command=lambda:data_arrange(4))
    if l > 3:
        treeview.heading(db_col[3], command=lambda:data_arrange(3))
    if l > 2:
        treeview.heading(db_col[2], command=lambda:data_arrange(2))
    if l > 1:
        treeview.heading(db_col[1], command=lambda:data_arrange(1))
    if l > 0:
        treeview.heading(db_col[0], command=lambda:data_arrange(0))

    scrollbar.config(command=treeview.yview)

    switch_frame(page2)

def yield_value():
    global lst, data_idx, db_list, db_col
    if len(data_lst) == 0:
        msgbox.showwarning("경고", "열 제목을 선택하십시오.")
        return
    lst = []
    for i in db_list:
        temp_lst = []
        for j in data_idx_lst:
            temp_lst.append(i[j])
        lst.append(temp_lst)
    data_idx = 0
    make_excel()

def make_excel():
    wb = Workbook()
    ws = wb.active
    l, s = len(data_lst), 0
    if 4 < l < 10:
        s = 1
    elif 2 < l < 5:
        s = 2
    elif l == 2:
        s = 3
    else:
        s = 5
    if s == 0:
        pass
    row, col, s_idx, p = 1, 1, 1, 1
    for lst_idx, i in enumerate(lst):
        for idx, val in enumerate(data_lst):
            ws.cell(row=row, column=idx+col).value = val
            ws.cell(row=row, column=idx+col).border = thin_border
            ws.cell(row=row, column=idx+col).alignment = Alignment(horizontal='center',vertical='center')
            ws.cell(row=row+1, column=idx+col).value = i[idx]
            ws.cell(row=row+1, column=idx+col).border = thin_border
            ws.cell(row=row+1, column=idx+col).alignment = Alignment(horizontal='center',vertical='center')
        if s_idx < s:
            s_idx += 1
            col += l + 1
        else:
            s_idx = 1
            col = 1
            row += 3
        if row == p*40:
            p += 1
            row += 1
        p_var.set((lst_idx+1)/len(lst)*100)
        progress_bar.update()
    try:     
        wb.save(save_path)
    except Exception:
        msgbox.showerror("에러", "파일이 열려있습니다.\n파일을 닫고 다시 시도하십시오.")
        p_var.set(0)
        progress_bar.update()
        return
    del_print.config(state=NORMAL)
    check_not_print.config(state=NORMAL)
    msgbox.showinfo("알림", "파일이 성공적으로 생성되었습니다.")

def printer():
    os.startfile(save_path, "print")

def file_config():
    os.startfile(save_path)

def go_home():
    scrollbar.destroy()
    treeview.destroy()
    data_erase_all()
    p_var.set(0)
    del_print.config(state=DISABLED)
    check_not_print.config(state=DISABLED)
    switch_frame(page1)

# 데이터 관리
#######################
page1 = Frame(root)
page2 = Frame(root)
page3 = Frame(root)
page4 = Frame(root)

db = ''
db_list = []
db_col = []
data_idx = 0
units = []
lst = []
save_path = resource_path('project_file/시험점수 정리.xlsx')
#######################

# 도움말 데이터
#######################
img_idx = 0

img_lst = []

img1 = PhotoImage(file=resource_path("project_file/1.png"))
img2 = PhotoImage(file=resource_path("project_file/2.png"))
img3 = PhotoImage(file=resource_path("project_file/3.png"))
img4 = PhotoImage(file=resource_path("project_file/4.png"))

img_lst.append(img1)
img_lst.append(img2)
img_lst.append(img3)
img_lst.append(img4)

txt_lst = []

txt1 = "1. 찾아보기 버튼을 누르고 사용할 엑셀파일을 찾으십시오."
txt2 = "2. 사용할 sheet의 이름을 선택하십시오."
txt3 = "3. 넣을 값에 해당하는 열 제목을 클릭하십시오.\n누른 순서대로 출력이 되고, 잘못 선택하였을 시 지울수 있습니다.\n값이 비어있었다면 \'Unnamed: 0\' 꼴로 값이 들어가는데,\n이는 그대로 출력 되므로 본래의 파일에서 임의로 수정해야 합니다."
txt4 = "4. 시작하기 버튼을 누르면 파일이 생성되고\n아래 두 버튼이 활성화 되어 원하는 옵션을 선택하십시오."

txt_lst.append(txt1)
txt_lst.append(txt2)
txt_lst.append(txt3)
txt_lst.append(txt4)
#######################

switch_frame(page1)

thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=(Side(style='thin')))

# page 1
# 파일 선택 프레임
select_frame = LabelFrame(page1, text="파일위치")
select_frame.pack(fill="both", padx=5, pady=5, ipady=5)

txt_file_dest = Entry(select_frame, state="readonly", readonlybackground="white")
txt_file_dest.pack(side="left", fill="x", expand=True, padx=5, pady=5, ipady=4)

btn_file_dest = Button(select_frame, text="찾아보기", width=10,command=open_file)
btn_file_dest.pack(fill="x", side="right", padx=5, pady=5)

# 데이터 선택 프레임
data_frame = Frame(page1)
data_frame.pack(padx=5, pady=5, ipady=5)

# sheet 선택
Label(data_frame, text="sheet 이름", width=8).pack(side="left", padx=5, pady=5)

sheet_box = ttk.Combobox(data_frame, height=10, width=10, state="readonly")
sheet_box.set("--파일없음--")
sheet_box.pack(side="right", padx=5, pady=5)

# 프레임전환 프레임
choose_frame = Frame(page1)
choose_frame.pack(fill="x", padx=5, pady=5, ipady=5)

help_btn = switch_button(choose_frame, "도움말", need_help)
help_btn.pack(side="left")

check_btn = switch_button(choose_frame, "다음", check)
check_btn.pack(side="right")

# page 2
# 돌아가기 프레임
search_frame = Frame(page2)
search_frame.pack(fill="x", expand=True, padx=5, pady=5, ipady=5)

return_btn = Button(search_frame, text="처음으로", command=go_home)
return_btn.pack(side="left")

# 자료 프레임
whole_data = Frame(page2)
whole_data.pack(fill="x", expand=True, padx=5, pady=5, ipady=5)

scrollbar = Scrollbar(whole_data)
treeview = ttk.Treeview(whole_data)

# 선택된 자료 출력 프레임
selected_data_frame = Frame(page2)
selected_data_frame.pack()

data_lst = []
data_idx_lst = []

data_selected = "선택된 열 : "
data_label = Label(selected_data_frame, text=data_selected)
data_label.pack(anchor="center")

btn_frame = Frame(selected_data_frame)
btn_frame.pack()
Button(btn_frame, text="하나\n지우기", command=data_erase).pack(side="left")
Button(btn_frame, text="모두\n지우기", command=data_erase_all).pack(side="right")

# 실행전환 프레임
choose_frame = Frame(page2)
choose_frame.pack(fill="x", padx=5, pady=5, ipady=5)

start_btn = switch_button(choose_frame, "시작하기", yield_value)
start_btn.pack(anchor="center")

# 진행 상황 Progress Bar
frame_progress = LabelFrame(page2, text="진행상황")
frame_progress.pack(fill="x", padx=5, pady=5, ipady=5)

p_var = DoubleVar()
progress_bar = ttk.Progressbar(frame_progress, maximum=100, variable=p_var)
progress_bar.pack(fill="x", padx=5, pady=5)

# 삭제여부 프레임
choose_delete = Frame(page2)
choose_delete.pack(fill="x", padx=5, pady=5, ipady=5)

del_print = switch_button(choose_delete, "바로\n출력하기", printer)
del_print.config(state=DISABLED)
del_print.pack(side="right", padx=5, pady=5)

check_not_print = switch_button(choose_delete, "파일\n확인하기", file_config)
check_not_print.config(state=DISABLED)
check_not_print.pack(side="right", padx=5, pady=5)

exit_btn = switch_button(choose_delete, "닫기", root.quit)
exit_btn.config(pady=12)
exit_btn.pack(side="left", padx=5, pady=5)

root.iconbitmap(resource_path("project_file/kirby.ico"))

root.resizable(False, False)
root.mainloop()