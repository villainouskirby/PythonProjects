import pandas as pd
from openpyxl import *
from openpyxl.styles import Alignment
from openpyxl.styles.borders import Border, Side

wb = Workbook()
ws = wb.active

thin_border = Border(left=Side(style='thin'),right=Side(style='thin'),top=Side(style='thin'),bottom=(Side(style='thin')))

lst = [[1, '김길동', 88, 100], [2, '김길순', 100, 92], [3, '김길자', 92, 88], [4, '김길식', 60, 60], [5, '김길수', 56, 68], [6, 0, 0, 0], [7, 0, 0, 0], [8, 0, 0, 0], [9, 0, 0, 0], [10, 0, 0, 0], [11, 0, 0, 0], [12, 0, 0, 0], [13, 0, 0, 0], [14, 0, 0, 0], [15, 0, 0, 0], [16, 0, 0, 0], [17, 0, 0, 0], [18, 0, 0, 0], [19, 0, 0, 0], [20, 0, 0, 0], [21, 0, 0, 0], [22, 0, 0, 0], [23, 0, 0, 0], [24, 0, 0, 0], [25, 0, 0, 0], [26, 0, 0, 0], [27, 0, 0, 0], [28, 0, 0, 0], [29, 0, 0, 0]]

lst2 = ['학번', '이름', '1과']
l, s = len(lst2), 0
if 4 < l < 10:
    s = 1
elif 2 < l < 5:
    s = 2
elif l == 2:
    s = 3
else:
    s = 5
if s == 0:
    pass
row, col, s_idx, p = 1, 1, 1, 1
for i in lst:
    for idx, val in enumerate(lst2):
        ws.cell(row=row, column=idx+col).value = val
        ws.cell(row=row, column=idx+col).border = thin_border
        ws.cell(row=row, column=idx+col).alignment = Alignment(horizontal='center',vertical='center')
        ws.cell(row=row+1, column=idx+col).value = i[idx]
        ws.cell(row=row+1, column=idx+col).border = thin_border
        ws.cell(row=row+1, column=idx+col).alignment = Alignment(horizontal='center',vertical='center')
    if s_idx < s:
        s_idx += 1
        col += l + 1
    else:
        s_idx = 1
        col = 1
        row += 3
    if row == p*40:
        p += 1
        row += 1

        
wb.save('project2/dex.xlsx')

